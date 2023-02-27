import openpyxl
import os

class excel:
    def __init__(self, file_name, sheet_format=['Data', 'Time', 'Enters', 'Exists', 'Inside']):
        """
        create excel file if not exist
        or open it if it exist
        """
        if not os.path.isdir('../Data'):
            os.mkdir('../Data')
        self.file_path = '../Data/{}.xlsx'.format(file_name)
        if os.path.exists(self.file_path):
            self.wb = openpyxl.load_workbook(self.file_path)
        else:
            self.wb = openpyxl.Workbook()
            self.wb.remove_sheet(self.wb['Sheet'])
            self.wb.create_sheet('Main')
            self.wb['Main'].append(sheet_format)
            self.wb.save(self.file_path)
    
    def add_sheet(self, sheet_name, sheet_format):
        """
        create new sheet
        with format [day, hour, in, out, ...etc]
        """
        if sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(sheet_name)
            self.wb[sheet_name].append(sheet_format)
            self.wb.save(self.file_path)

    def append_to_sheet(self, sheet_name, line):
        """
        to update a sheet with new lines
        """
        self.wb[sheet_name].append(line)
        self.wb.save(self.file_path)

    def close(self):
        self.wb.close()
